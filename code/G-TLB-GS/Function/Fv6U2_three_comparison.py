import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime, timedelta
import os
from matplotlib.animation import FuncAnimation
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
import warnings
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from source.Fv6_generate_time_ticks import Fv6_generate_time_ticks


def Fv6U2_three_comparison(site_batch1_data, site_predictions, site_predict_timestamps_matrix,
                           site_batch2_data, site_params, output_prediction_folder, title, ylabel,
                           time_interval_hours, rotation_user,batch3_data_folder_path):
    # batch3_data_folder_path：ERA5文件夹，用于SHAP分析结果命名

    # 0.0获取文件夹中的所有文件名（去除后缀），并按升序排序
    file_names = []
    if os.path.isdir(batch3_data_folder_path):
        for filename in os.listdir(batch3_data_folder_path):
            if os.path.isfile(os.path.join(batch3_data_folder_path, filename)):
                # 去除文件后缀
                base_name = os.path.splitext(filename)[0]
                file_names.append(base_name)

    # 按文件名升序排序
    file_names.sort()

    # 0. 全局设置 Times New Roman 字体
    plt.rcParams['font.family'] = 'Times New Roman'
    plt.rcParams['mathtext.fontset'] = 'stix'
    warnings.filterwarnings('ignore', category=UserWarning)

    # 创建输出文件夹
    animation_output_folder = os.path.join(output_prediction_folder)
    os.makedirs(animation_output_folder, exist_ok=True)

    # 1. 预处理：确保所有时间戳都是数值型（UNIX时间戳）
    def ensure_timestamps_numeric(data_list):
        for idx, data in enumerate(data_list):
            if data is not None and len(data) > 0:
                if not np.issubdtype(data[:, 0].dtype, np.number):
                    data[:, 0] = data[:, 0].astype(float)
        return data_list

    site_batch1_data = ensure_timestamps_numeric(site_batch1_data)
    if site_batch2_data is not None:
        site_batch2_data = ensure_timestamps_numeric(site_batch2_data)

    # 2. 准备评估指标数据结构
    site_metrics = {}
    global_metrics = []
    global_mae_list = []
    global_rmse_list = []
    global_r2_list = []

    # 3. 遍历每个站点
    for site_idx in range(len(site_params)):
        try:
            print(f"\n处理站点 {site_idx + 1}/{len(site_params)}...")

            # 获取当前站点的文件名（按升序排序）
            if site_idx < len(file_names):
                file_title = file_names[site_idx]
            else:
                # 如果站点数多于文件名数，循环使用文件名
                file_title = file_names[site_idx % len(file_names)]

            # 获取当前站点的数据
            batch1_data = site_batch1_data[site_idx]
            predictions = site_predictions.get(site_idx, None)
            predict_ts_matrix = site_predict_timestamps_matrix.get(site_idx, None)
            batch2_data = site_batch2_data[site_idx] if site_batch2_data else None

            # 从site_params获取站点参数
            site_param = site_params[site_idx]
            use_forward_hours = site_param["use_forward_hours"]
            use_backward_hours = site_param["use_backward_hours"]

            # 4. 计算当前站点的时间戳范围
            if predict_ts_matrix is None or predict_ts_matrix.size == 0:
                print(f"站点 {site_idx + 1} 没有预测数据，跳过")
                continue

            min_pred_time = np.min(predict_ts_matrix)
            max_pred_time = np.max(predict_ts_matrix)
            site_min_ts = min_pred_time - use_forward_hours * 3600  # 考虑历史时间步
            site_max_ts = max_pred_time
            print(f"站点 {site_idx + 1} 时间从: {datetime.fromtimestamp(site_min_ts)} 到 {datetime.fromtimestamp(site_max_ts)}")

            # 5. 准备绘图数据
            b1_timestamps = batch1_data[:, 0] if batch1_data is not None else np.array([])
            b1_values = batch1_data[:, 1] if batch1_data is not None else np.array([])

            # 6. 创建基础图形
            fig, ax = plt.subplots(figsize=(15, 4))

            # 7. 绘制历史背景数据
            if b1_timestamps.size > 0:
                mask = (b1_timestamps >= site_min_ts) & (b1_timestamps <= site_max_ts)
                ax.plot(b1_timestamps[mask], b1_values[mask], color='#f9c00c', alpha=0.5, linewidth=2, label='Forecast')

            # 8. 绘制实测数据背景
            if batch2_data is not None and len(batch2_data) > 0:
                b2_timestamps = batch2_data[:, 0]
                b2_values = batch2_data[:, 1]
                mask = (b2_timestamps >= site_min_ts) & (b2_timestamps <= site_max_ts)
                ax.plot(b2_timestamps[mask], b2_values[mask], color='#7200da', alpha=0.5, linewidth=2, label='Observed')

            # 9. 设置标题和标签
            history_steps = site_param["use_forward_hours"]
            future_steps = site_param["use_backward_hours"]
            plt.ylabel(ylabel)
            plt.legend(loc='upper left')

            # 10. 设置时间范围
            ax.set_xlim(site_min_ts, site_max_ts)

            # 11. 设置X轴刻度
            tick_locations, tick_labels = Fv6_generate_time_ticks(site_min_ts, site_max_ts, time_interval_hours)
            ax.set_xticks(tick_locations)
            ax.set_xticklabels(tick_labels, rotation=rotation_user)
            ax.grid(True, linestyle=':', alpha=0.3)

            # 12. 确保预测数据存在
            num_frames = predict_ts_matrix.shape[0]
            if num_frames == 0:
                print(f"站点 {site_idx + 1} 没有预测帧，跳过动画")
                plt.close(fig)
                continue

            # 14. 创建动画元素
            history_line, = ax.plot([], [], color='#00b9f1', linewidth=3, alpha=1, label='DL_Forecast')
            ax.legend(loc='upper left', fontsize=10)
            prediction_line, = ax.plot([], [], color='#00b9f1', linewidth=3, alpha=1)
            current_point = ax.scatter([], [], c='#00b9f1', s=40, alpha=1, edgecolors='black', zorder=10)

            # 15. 创建动画更新函数
            def update(frame):
                # 获取当前预测点的相关信息
                current_time = predict_ts_matrix[frame, 0]  # 当前时刻（预测起始时间）

                # 更新标题显示当前时刻
                time_str = datetime.utcfromtimestamp(current_time).strftime('%Y-%m-%d %H:%M:%S')
                ax.set_title(f"-{history_steps}h+{future_steps}h / {title}{file_title}\nCurrent Time: {time_str}", fontsize=14)

                # 获取历史数据序列（当前时刻之前的数据）
                if b1_timestamps.size > 0:
                    # 计算历史时间窗口的起始时间
                    history_start = current_time - use_forward_hours * 3600     # 考虑历史时间步

                    # 获取用于当前预测的所用历史数据点
                    history_mask = (b1_timestamps >= history_start) & (b1_timestamps <= current_time)

                    # 确保至少有一个数据点
                    if np.any(history_mask):
                        history_line.set_data(b1_timestamps[history_mask], b1_values[history_mask])
                    else:
                        # 如果没有历史数据，设置空数据
                        history_line.set_data([], [])

                # 获取预测序列
                pred_times = predict_ts_matrix[frame]  # 输出层预测时间戳
                pred_values = predictions[frame]  # 值

                # 更新预测曲线 - 确保有数据
                if len(pred_times) > 0 and len(pred_values) > 0:
                    prediction_line.set_data(pred_times, pred_values)
                else:
                    prediction_line.set_data([], [])

                # 更新当前时刻点 - 确保有数据
                if b1_timestamps.size > 0:
                    # 在当前时间位置添加一个点
                    current_value = b1_values[b1_timestamps == current_time]
                    if len(current_value) == 0:  # 如果当前时间点不在输入序列中
                        # 尝试从预测序列中获取值
                        if len(pred_values) > 0:
                            current_value = [pred_values[0]]
                        else:
                            current_value = [0]
                    current_point.set_offsets([[current_time, current_value[0]]])   # 当前值取预测值
                else:
                    # 如果没有输入数据，使用预测值
                    if len(pred_values) > 0:
                        current_point.set_offsets([[current_time, pred_values[0]]])
                    else:
                        current_point.set_offsets([[current_time, 0]])

                # 返回所有更新后的对象
                return history_line, prediction_line, current_point

            # 16. 收集整点时刻帧
            whole_hour_frames = []
            for frame in range(num_frames):
                current_time = predict_ts_matrix[frame, 0]
                dt = datetime.utcfromtimestamp(current_time)
                if dt.minute == 0 and dt.second == 0:
                    whole_hour_frames.append(frame)

            if not whole_hour_frames:
                print(f"站点 {site_idx + 1} 没有整小时的预测帧，跳过动画")
                plt.close(fig)
                continue

            max_frames = 1000   # 其实全绘制即可
            if len(whole_hour_frames) > max_frames:
                skip = max(1, len(whole_hour_frames) // max_frames)
                frames_to_animate = whole_hour_frames[::skip]
            else:
                frames_to_animate = whole_hour_frames

            animation = FuncAnimation(
                fig, update,
                frames=frames_to_animate,
                interval=100,
                blit=True
            )

            # 17. 保存动画为GIF
            gif_filename = f"{file_title}_forecast_animation.gif"
            gif_path = os.path.join(animation_output_folder, gif_filename)
            animation.save(gif_path, writer='pillow', fps=10)   # 设置帧率为每秒 10 帧
            print(f"站点 {site_idx + 1} .gif文件保存至: {gif_path}")

            # 保存整点时刻静态图片
            whole_hour_image_dir = os.path.join(output_prediction_folder, f'site_{site_idx + 1}_whole_hour_images')
            os.makedirs(whole_hour_image_dir, exist_ok=True)

            # 每隔720个整点帧的保存静态图片
            save_interval = 720  # 每save_interval个帧保存一张图片
            for i, frame in enumerate(frames_to_animate):
                # 只保存每24帧中的第一帧
                if i % save_interval == 0:
                    update(frame)
                    current_time = predict_ts_matrix[frame, 0]
                    time_str = datetime.fromtimestamp(current_time).strftime('%Y%m%d_%H%M%S')   # 无需UTC转换
                    img_filename = f'frame_{frame}_{time_str}.png'
                    img_path = os.path.join(whole_hour_image_dir, img_filename)
                    fig.savefig(img_path, dpi=300, bbox_inches='tight')
                    # print(f"站点 {site_idx + 1} 的整点时刻帧 {frame} 已保存为: {img_path}")

            plt.close(fig)

            # 18. 计算评估指标
            if batch2_data is not None and len(batch2_data) > 0:
                b2_timestamps = batch2_data[:, 0]
                b2_values = batch2_data[:, 1]

                # 初始化站点指标
                site_metrics[site_idx] = {'all_frames': []}
                site_mae_list = []
                site_rmse_list = []
                site_r2_list = []

                # 计算所有样本的评价指标
                for frame in range(num_frames):
                    pred_times = predict_ts_matrix[frame]   # 输出层预测数据时间戳
                    pred_values = predictions[frame]

                    observed_mask = np.isin(b2_timestamps, pred_times)  # 创建一个布尔数组 observed_mask，对于 b2_timestamps 中的每个时间戳，检查它是否存在于 pred_times 中
                    if not np.any(observed_mask):   # 检查 observed_mask 中是否有至少一个 True 值
                        continue

                    observed_values = np.array(b2_values)[observed_mask]    # 观测值
                    pred_dict = {float(t): p for t, p in zip(pred_times, pred_values)}  # DL模型预测值
                    matched_pred_values = [pred_dict[t] for t in b2_timestamps[observed_mask]]  # 在

                    if len(observed_values) > 1:
                        mae = mean_absolute_error(observed_values, matched_pred_values)
                        rmse = np.sqrt(mean_squared_error(observed_values, matched_pred_values))
                        # 在R²计算前添加以下代码
                        if np.var(observed_values) < 1e-3:  # 判断实测数据变化是否过小
                            r2 = np.nan  # 变化太小则忽略R²
                        else:
                            r2 = r2_score(observed_values, matched_pred_values)
                    elif len(observed_values) == 1:
                        mae = np.abs(observed_values[0] - matched_pred_values[0])
                        rmse = mae
                        r2 = np.nan
                    else:
                        continue

                    # 添加到列表
                    site_mae_list.append(mae)
                    site_rmse_list.append(rmse)
                    site_r2_list.append(r2)
                    global_mae_list.append(mae)
                    global_rmse_list.append(rmse)
                    global_r2_list.append(r2)

                    # 存储详细数据
                    pred_time_str = datetime.fromtimestamp(pred_times[0]).strftime('%Y-%m-%d %H:%M:%S')
                    site_metrics[site_idx]['all_frames'].append({
                        'Frame': frame,
                        'ForecastTime': pred_time_str,
                        'MAE': mae,
                        'RMSE': rmse,
                        'R2': r2
                    })

                # 计算站点平均指标
                if site_mae_list:
                    avg_mae = np.mean(site_mae_list)
                    avg_rmse = np.mean(site_rmse_list)
                    avg_r2 = np.nanmean(site_r2_list)

                    global_metrics.append({
                        'Site': site_idx,
                        'Avg_MAE': avg_mae,
                        'Avg_RMSE': avg_rmse,
                        'Avg_R2': avg_r2,
                        'Num_Frames': num_frames
                    })

        except Exception as e:
            import traceback
            print(f"Site {site_idx} processing failed: {str(e)}")
            traceback.print_exc()
            plt.close('all')

    # 19. 保存评估指标到Excel
    if site_metrics:
        # 创建Excel工作簿
        wb = Workbook()

        # 添加站点详细数据工作表
        for site_idx, metrics in site_metrics.items():
            all_frames_df = pd.DataFrame(metrics['all_frames'])
            ws_all = wb.create_sheet(title=f"Site_{site_idx + 1}")
            for r in dataframe_to_rows(all_frames_df, index=False, header=True):
                ws_all.append(r)

        # 添加站点汇总工作表
        if global_metrics:
            site_summary_df = pd.DataFrame(global_metrics)
            ws_summary = wb.create_sheet(title="Summary")
            for r in dataframe_to_rows(site_summary_df, index=False, header=True):
                ws_summary.append(r)

            # 添加全局汇总
            if global_mae_list:
                global_avg_mae = np.mean(global_mae_list)
                global_avg_rmse = np.mean(global_rmse_list)
                global_avg_r2 = np.nanmean(global_r2_list)

                ws_summary.append([])
                ws_summary.append(["Global Summary"])
                ws_summary.append(["Avg_MAE", "Avg_RMSE", "Avg_R2"])
                ws_summary.append([global_avg_mae, global_avg_rmse, global_avg_r2])

        # 删除默认工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 保存Excel文件
        excel_path = os.path.join(output_prediction_folder, 'forecast_metrics.xlsx')
        wb.save(excel_path)
        print(f"\n预测指标已保存到Excel文件: {excel_path}")
    else:
        print("\n警告：没有观测到的可用数据，未生成指标")

    print("所有站点处理成功")